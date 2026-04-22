import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from collections import defaultdict

SKIP_TAG_PREFIXES = {
    "Forney_Atmos",
    "Kendall_Fuel",
    "Blackstone_BLA1_AGC_SP_CONTROL",
    "Forney_Kinder"
}

PLANTS = {
    "Baldwin",
    "Bellingham",
    "Blackstone",
    "Calumet",
    "Casco Bay",
    "Decordova",
    "Ennis",
    "Fayette",
    "Forney",
    "Graham",
    "Hanging Rock",
    "Hays",
    "Independence",
    "Kendall",
    "Kincaid",
    "Lake Hubbard",
    "Lake Road",
    "Lamar",
    "Liberty",
    "Martin Lake",
    "Masspower",
    "Miami Fort",
    "Midlothian",
    "Milford",
    "Morgan Creek",
    "Moss Landing",
    "Newton",
    "Oak Grove",
    "Odessa",
    "Ontelaunee",
    "Permian Basin",
    "Pleasants",
    "Stryker Creek",
    "Trinidad",
    "Washington",
    "Wise"
}

def write_output(map):
    wb = Workbook() #create a new workbook
    default_sheet = wb.active
    wb.remove(default_sheet)    #get rid of the default sheet that comes with creating a new workbook

    for plant, taglist in sorted(map.items()):
        ws = wb.create_sheet(title = f"{plant}_Create")
        ws.append(['Create Tag'])   #create header column
        ws["A1"].font = Font(bold=True)    #make header column bold

        row = 2
        for tag in taglist:
            ws.cell(row=row, column=1, value=tag)
            row += 1

    wb.save("fleet_performance_tags_to_create.xlsx")

#helper function to extract only the plant name from the tag. E.g. if given tag is Hays 1_Gas Turbine_XYZ this function will extract 'Hays'
def extract_plant_name(tagname):
    if pd.isna(tagname):
        return None

    #remove leading & trailing whitespace
    tagname = str(tagname).strip()

    #skip over the tags in our global 'skip' set
    for skip_prefix in SKIP_TAG_PREFIXES:
        if tagname == skip_prefix or tagname.startswith(skip_prefix + " "):
            return None

    #sort plants so plants with longer names are first in the list. e.g. if our tag was Lake Road_Fuel 'startwith()' will pick up 'Lake Road' first before 'Lake '
    #which is required because plants like lake hubbard and lake road exist thus only grabbing "lake" introduces ambiguity. sorting gets rid of this issue.
    for plant in sorted(PLANTS, key=len, reverse=True):
        if (tagname == plant or tagname.startswith(plant + " ") or tagname.startswith(plant + "_")):
            return plant
    return tagname.split(" ", 1)[0]

def main():
    print("Starting main program...\n")

    #get cwd so that we can grab the excel file with our data
    cwd = os.getcwd()
    file_path = f"{cwd}\\pitag_list.xlsx"

    #read the excel file into a dataframe and initialize a map 
    df = pd.read_excel(file_path)
    plant_tag_mappings = defaultdict(list)

    for pitag in df["Create these POCPI tags"]:
        plant = extract_plant_name(pitag)
        if plant is None:
            continue
        plant_tag_mappings[plant].append(pitag)

    # print(plant_tag_mappings.keys())
    # print("\n")
    # print(plant_tag_mappings)

    #send the plant to pitag mappings to a new excel file to be written
    write_output(plant_tag_mappings)

    print("\n=======Program complete.=======\n")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Error running main - {e}")
